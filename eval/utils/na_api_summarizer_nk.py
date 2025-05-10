# pip install pandas anthropic openpyxl

import os
import re
import glob
import pandas as pd
import anthropic
import time
from tqdm import tqdm
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Configuration
API_KEY = #"YOUR_API_KEY_HERE"  # Replace with your actual API key
INPUT_DIR = "./"  # Directory containing text files (current directory)
OUTPUT_FILE = "news_summaries.xlsx"  # Output Excel file
MAX_SUMMARY_WORDS = 200  # Maximum words for the summary

# Rate limiting configuration
TOKEN_LIMIT_PER_MINUTE = 20000  # Claude's rate limit
WORD_TO_TOKEN_RATIO = 1.75  # Assume 1 word = 1.75 tokens
MAX_DOCUMENT_TOKENS = 20000  # Maximum tokens for document processing

# Initialize Claude client
client = anthropic.Anthropic(api_key=API_KEY)

# Token usage tracker
class TokenUsageTracker:
    def __init__(self, limit_per_minute):
        self.limit = limit_per_minute
        self.used = 0
        self.last_reset = time.time()
    
    def estimate_tokens(self, content):
        """Estimate token count based on word count"""
        word_count = len(content.split())
        return int(word_count * WORD_TO_TOKEN_RATIO)
    
    def can_process(self, content):
        """Check if we can process the content without exceeding rate limits"""
        # Reset counter if more than a minute has passed
        current_time = time.time()
        if current_time - self.last_reset >= 60:
            self.used = 0
            self.last_reset = current_time
        
        # Estimate tokens for this content
        estimated_tokens = self.estimate_tokens(content)
        
        # Check if we have enough quota
        return self.used + estimated_tokens <= self.limit
    
    def record_usage(self, content):
        """Record token usage after processing content"""
        self.used += self.estimate_tokens(content)
    
    def wait_for_reset(self):
        """Wait until the rate limit resets - always wait a full minute"""
        print(f"Rate limit approaching. Waiting for 60 seconds...")
        time.sleep(60)  # Always wait a full minute
        
        # Reset counter
        self.used = 0
        self.last_reset = time.time()

# Initialize token tracker
token_tracker = TokenUsageTracker(TOKEN_LIMIT_PER_MINUTE)

def truncate_to_token_limit(text, max_tokens=MAX_DOCUMENT_TOKENS):
    """
    Truncate text to stay within token limit
    
    Args:
        text: The text to truncate
        max_tokens: Maximum tokens allowed
        
    Returns:
        Truncated text and original word count
    """
    max_words = int(max_tokens / WORD_TO_TOKEN_RATIO)
    words = text.split()
    original_word_count = len(words)
    
    if len(words) <= max_words:
        return text, original_word_count
    
    print(f"Truncating document from {len(words)} words to {max_words} words to stay within token limit")
    return " ".join(words[:max_words]), original_word_count

def extract_metadata(text):
    """Extract title, date, and URL from text content"""
    metadata = {}
    
    # Extract metadata using regex patterns with flexible newlines
    title_match = re.search(r'Title:\s*(.*?)(?:\r?\n|$)', text, re.IGNORECASE)
    date_match = re.search(r'Date:\s*(.*?)(?:\r?\n|$)', text, re.IGNORECASE)
    url_match = re.search(r'URL:\s*(.*?)(?:\r?\n|$)', text, re.IGNORECASE)
    source_match = re.search(r'Source:\s*(.*?)(?:\r?\n|$)', text, re.IGNORECASE)
    topic_match = re.search(r'Topic:\s*(.*?)(?:\r?\n|$)', text, re.IGNORECASE)
    category_match = re.search(r'Category:\s*(.*?)(?:\r?\n|$)', text, re.IGNORECASE)
    
    # Store extracted metadata
    metadata['title'] = title_match.group(1).strip() if title_match else ""
    metadata['date'] = date_match.group(1).strip() if date_match else ""
    metadata['url'] = url_match.group(1).strip() if url_match else ""
    metadata['source'] = source_match.group(1).strip() if source_match else ""
    metadata['topic'] = topic_match.group(1).strip() if topic_match else ""
    metadata['category'] = category_match.group(1).strip() if category_match else ""
    
    return metadata

def extract_content(text):
    """Extract the main content from the text file"""
    # Split the text into lines, handling both Unix and Windows line endings
    lines = re.split(r'\r?\n', text)
    
    # Find the line with journalist info
    journalist_line = -1
    for i, line in enumerate(lines):
        if 'Journalist:' in line:
            journalist_line = i
            break
    
    if journalist_line == -1:
        # If we can't find journalist line, look for the end of metadata section
        for i, line in enumerate(lines):
            if i > 5 and line.strip() == '' and i+1 < len(lines) and lines[i+1].strip() != '':
                journalist_line = i
                break
    
    if journalist_line == -1:
        # Fallback: just use the first 10 lines as metadata
        journalist_line = 10
    
    # Skip empty lines after metadata section
    content_start = journalist_line + 1
    while content_start < len(lines) and not lines[content_start].strip():
        content_start += 1
    
    # Get the actual content
    content_lines = lines[content_start:]
    content = '\n'.join(content_lines)
    
    # Remove additional articles at the end (they often start with ##)
    hash_index = content.find('\n##')
    if hash_index != -1:
        content = content[:hash_index].strip()
    
    return content.strip()

def generate_summary(content, metadata):
    """Generate a summary using Claude 3.7 API"""
    
    # Prepare a rich context for the API
    title = metadata.get('title', '')
    source = metadata.get('source', '')
    date = metadata.get('date', '')
    topic = metadata.get('topic', '')
    category = metadata.get('category', '')
    
    context = f"Title: {title}\n"
    if source:
        context += f"Source: {source}\n"
    if date:
        context += f"Date: {date}\n"
    if topic:
        context += f"Topic: {topic}\n"
    if category:
        context += f"Category: {category}\n"
    
    # Modified prompt to explicitly avoid titles and # symbols
    prompt = f"""Summarize the following news article by extracting and paraphrasing only the core information and key facts in a single paragraph, limited to 60 words.
            Strict formatting rules:
            - IMPORTANT: Do NOT generate the summary with any titles or headings
            - Do not include rhetorical questions.
            - Remove all conversational or reflective phrases (e.g., "Let's break down...", "Do you want to...").
            - Do not include any introductory or closing remarks.
            - Write in a neutral, journalistic tone.
            - Present all key facts and information clearly and concisely in paragraph form (no bullet points).
            - Ensure that the summary always ends with a full and complete sentence.
            - Keep the summary short, recommended to be around 100 words.
            - Begin directly with the summary text without any prefixes.

            {context}

            {content}"""
    
    # Check if we can make this API call without hitting rate limits
    if not token_tracker.can_process(content):
        token_tracker.wait_for_reset()
    
    try:
        response = client.messages.create(
            model="claude-3-7-sonnet-20250219",
            max_tokens=1024,
            temperature=0.0,
            system="You are a professional news summarizer. Create concise, accurate summaries that capture the essential information of news articles. Your summaries should be factual, well-structured, and contain the key points from the original article. Always maintain the tone and perspective of the original article. NEVER start summaries with titles, headings, or # symbols.",
            messages=[
                {"role": "user", "content": prompt}
            ]
        )
        
        # Record token usage
        token_tracker.record_usage(content)
        
        summary = response.content[0].text
        
        # Further clean the summary to remove any remaining '#' symbols or titles at the beginning
        summary = re.sub(r'^#\s*[^\n]*\n', '', summary)  # Remove title with # if present
        summary = re.sub(r'^Title:', '', summary)  # Remove "Title:" prefix if present
        summary = re.sub(r'^Summary:', '', summary)  # Remove "Summary:" prefix if present
        summary = summary.strip()
        
        # Ensure summary is under MAX_SUMMARY_WORDS words
        words = summary.split()
        if len(words) > MAX_SUMMARY_WORDS:
            summary = " ".join(words[:MAX_SUMMARY_WORDS]) + "..."
            
        return summary
        
    except Exception as e:
        print(f"Error generating summary: {e}")
        return f"Error generating summary: {e}"

def count_words(text):
    """Count the number of words in text"""
    return len(text.split())

def read_existing_excel():
    """Read the existing Excel file if it exists"""
    if not os.path.exists(OUTPUT_FILE):
        print(f"Excel file {OUTPUT_FILE} not found. Will create a new one.")
        return None
    
    try:
        df = pd.read_excel(OUTPUT_FILE)
        print(f"Successfully read {len(df)} entries from {OUTPUT_FILE}")
        return df
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

def needs_regeneration(summary):
    """Check if a summary needs to be regenerated"""
    # Check for NaN and None
    if not isinstance(summary, str):
        return True

    # Check for empty strings
    summary = summary.strip()
    if not summary:  # This will catch empty strings
        return True
        
    summary = summary.strip()
    return (summary.startswith('#') or 
            summary.startswith('Error') or
            summary.startswith('Title:') or
            summary.startswith('Summary:'))

def process_existing_excel():
    """Process entries in the existing Excel file"""
    # Read the existing Excel file
    df = read_existing_excel()
    if df is None or len(df) == 0:
        print("No existing data to process.")
        return False
    
    # Check if "Document Word Count" column exists, add it if it doesn't
    if 'Document Word Count' not in df.columns:
        df['Document Word Count'] = 0
        print("Added 'Document Word Count' column to the existing spreadsheet.")
    
    # Identify entries with problematic summaries (need regeneration)
    summary_mask = df['Summary'].apply(needs_regeneration)
    summary_count = summary_mask.sum()
    
    # Identify entries with zero word count 
    word_count_mask = df['Document Word Count'] == 0
    word_count_count = word_count_mask.sum()
    
    if summary_count == 0 and word_count_count == 0:
        print("No entries need processing.")
        return False
    
    print(f"Found {summary_count} summaries that need regeneration.")
    print(f"Found {word_count_count} entries with zero word count that need calculation.")
    
    # First, process entries that need only word count calculation (not summary regeneration)
    word_count_only_mask = word_count_mask & ~summary_mask
    word_count_only_count = word_count_only_mask.sum()
    
    if word_count_only_count > 0:
        print(f"Processing {word_count_only_count} entries that need word count calculation only...")
        
        for idx, row in tqdm(df[word_count_only_mask].iterrows(), total=word_count_only_count, desc="Calculating word counts"):
            file_path = os.path.join(INPUT_DIR, row['Filename'])
            
            if not os.path.exists(file_path):
                print(f"File not found: {file_path}. Skipping.")
                continue
            
            try:
                print(f"Calculating word count for: {row['Filename']}")
                
                # Read file content
                with open(file_path, 'r', encoding='utf-8', errors='replace') as file:
                    content = file.read()
                
                # Extract main content
                article_content = extract_content(content)
                
                # Calculate and update word count
                original_word_count = count_words(article_content)
                df.at[idx, 'Document Word Count'] = original_word_count
                
                print(f"  Updated word count: {original_word_count} words")
                
                # Save after each update to avoid losing progress
                df.to_excel(OUTPUT_FILE, index=False)
                
            except Exception as e:
                print(f"  Error calculating word count for {row['Filename']}: {e}")
    
    # Next, process entries that need summary regeneration
    if summary_count > 0:
        print(f"Processing {summary_count} entries that need summary regeneration...")
        
        for idx, row in tqdm(df[summary_mask].iterrows(), total=summary_count, desc="Regenerating summaries"):
            file_path = os.path.join(INPUT_DIR, row['Filename'])
            
            if not os.path.exists(file_path):
                print(f"File not found: {file_path}. Skipping.")
                continue
            
            try:
                print(f"Regenerating summary for: {row['Filename']}")
                
                # Read file content
                with open(file_path, 'r', encoding='utf-8', errors='replace') as file:
                    content = file.read()
                
                # Extract metadata and main content
                metadata = extract_metadata(content)
                article_content = extract_content(content)
                
                # Update word count if needed
                if row['Document Word Count'] == 0:
                    original_word_count = count_words(article_content)
                    df.at[idx, 'Document Word Count'] = original_word_count
                    print(f"  Updated word count: {original_word_count} words")
                
                # Truncate if needed to stay within token limits
                truncated_content, _ = truncate_to_token_limit(article_content)
                
                # Generate new summary
                new_summary = generate_summary(truncated_content, metadata)
                
                # Update the DataFrame
                df.at[idx, 'Summary'] = new_summary
                df.at[idx, 'Summary Length'] = count_words(new_summary)
                
                print(f"  Updated summary for {row['Filename']}")
                
                # Save after each update to avoid losing progress
                df.to_excel(OUTPUT_FILE, index=False)
                
            except Exception as e:
                print(f"  Error regenerating summary for {row['Filename']}: {e}")
    
    print("Processing of existing entries completed.")
    return True

def get_new_text_files(existing_df):
    """Get list of text files not already in the Excel file"""
    # Find all text files in the directory
    all_text_files = glob.glob(os.path.join(INPUT_DIR, "*.txt"))
    
    # If there's an existing dataframe, get the filenames already processed
    if existing_df is not None and len(existing_df) > 0:
        existing_filenames = set(existing_df['Filename'].tolist())
    else:
        existing_filenames = set()
    
    # Find new text files
    new_text_files = []
    for file_path in all_text_files:
        filename = os.path.basename(file_path)
        if filename not in existing_filenames:
            new_text_files.append(file_path)
    
    return new_text_files

def append_to_excel(new_data, existing_df):
    """Append new data to existing Excel file"""
    if existing_df is not None and len(existing_df) > 0:
        # Combine existing and new data
        all_data = pd.concat([existing_df, pd.DataFrame(new_data)], ignore_index=True)
    else:
        # Create new dataframe if no existing data
        all_data = pd.DataFrame(new_data)
    
    # Save to Excel with formatting
    save_to_excel(all_data)

def save_to_excel(df):
    """Save DataFrame to Excel with proper formatting"""
    # Save to Excel with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Article Summaries"
    
    # Add headers
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            if col_num == headers.index("Summary") + 1:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Auto-adjust column width
    for column in range(1, len(headers) + 1):
        if column == headers.index("Summary") + 1:
            # Make summary column wider
            ws.column_dimensions[get_column_letter(column)].width = 60
        else:
            # Auto-adjust other columns based on content
            max_length = 0
            for row in range(1, len(df) + 2):
                cell_value = str(ws.cell(row=row, column=column).value or "")
                max_length = max(max_length, len(cell_value))
            ws.column_dimensions[get_column_letter(column)].width = min(max_length + 2, 30)
    
    # Freeze the header row
    ws.freeze_panes = "A2"
    
    # Save the workbook
    wb.save(OUTPUT_FILE)
    print(f"\nSummaries saved to {OUTPUT_FILE}")

def process_files():
    """Process all text files and create/update Excel spreadsheet"""
    
    # First, process any existing Excel file entries
    if process_existing_excel():
        print("Processing of existing entries completed.")
    
    # Read the existing Excel file to get list of already processed files
    existing_df = read_existing_excel()
    
    # Get only the new text files that haven't been processed yet
    new_text_files = get_new_text_files(existing_df)
    
    if not new_text_files:
        print("\nNo new text files found to process.")
        return
    
    print(f"\nFound {len(new_text_files)} new text files to process.")
    
    # Prepare data for new files
    new_data = []
    
    for i, file_path in enumerate(tqdm(new_text_files, desc="Processing new text files")):
        filename = os.path.basename(file_path)
        print(f"[{i+1}/{len(new_text_files)}] Processing: {filename}")
        
        try:
            # Read file content
            with open(file_path, 'r', encoding='utf-8', errors='replace') as file:
                content = file.read()
            
            # Extract metadata and main content
            metadata = extract_metadata(content)
            article_content = extract_content(content)
            
            # Get original word count
            original_word_count = count_words(article_content)
            
            # Truncate if needed to stay within token limits
            truncated_content, _ = truncate_to_token_limit(article_content)
            
            # Log extracted content length
            content_word_count = count_words(truncated_content)
            print(f"  Extracted {content_word_count} words from article content (original: {original_word_count})")
            
            # Generate summary
            print(f"  Generating summary for: {metadata['title']}")
            summary = generate_summary(truncated_content, metadata)
            
            # Calculate summary length
            summary_length = count_words(summary)
            print(f"  Generated summary of {summary_length} words")
            
            # Format date if possible
            date_str = metadata['date']
            formatted_date = date_str
            try:
                # Try common date formats
                if re.match(r'\d{2}/\d{2}/\d{4}', date_str):
                    date_obj = datetime.strptime(date_str, '%d/%m/%Y')
                    formatted_date = date_obj.strftime('%Y-%m-%d')
                elif re.match(r'\d{1,2}/\d{1,2}/\d{4}', date_str):
                    date_obj = datetime.strptime(date_str, '%d/%m/%Y')
                    formatted_date = date_obj.strftime('%Y-%m-%d')
                elif re.match(r'\d{2}/\d{1,2}/\d{4}', date_str):
                    date_obj = datetime.strptime(date_str, '%d/%m/%Y')
                    formatted_date = date_obj.strftime('%Y-%m-%d')
            except Exception as date_error:
                print(f"  Warning: Could not parse date '{date_str}': {date_error}")
                # If date parsing fails, keep original string
                pass
            
            # Add to new data
            new_data.append({
                'Filename': filename,
                'Title': metadata['title'],
                'Date': formatted_date,
                'Source': metadata['source'],
                'Topic': metadata['topic'],
                'Category': metadata['category'],
                'Summary': summary,
                'Summary Length': summary_length,
                'Document Word Count': original_word_count,  # Add original word count
                'URL': metadata['url']
            })
            
        except Exception as e:
            print(f"  Error processing {filename}: {e}")
            import traceback
            traceback.print_exc()
            
    # Append new data to existing Excel file
    if new_data:
        append_to_excel(new_data, existing_df)
        print(f"\nCompleted! Added {len(new_data)} new summaries to {OUTPUT_FILE}")
    else:
        print("\nNo new data to save.")

if __name__ == "__main__":
    print("Starting text summarization process...")
    process_files()
    print("Process completed!")
