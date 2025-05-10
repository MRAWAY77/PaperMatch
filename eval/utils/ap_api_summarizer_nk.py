# pip install pandas anthropic openpyxl PyPDF2 pdfplumber tqdm nltk

import os
import re
import glob
import pandas as pd
import anthropic
import time
import PyPDF2
import pdfplumber
import io
import string
from tqdm import tqdm
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Import NLTK libraries for text processing (alternative to spaCy)
import nltk
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer

# Download necessary NLTK data (only needed first time)
try:
    nltk.data.find('tokenizers/punkt')
except LookupError:
    nltk.download('punkt')
try:
    nltk.data.find('corpora/stopwords')
except LookupError:
    nltk.download('stopwords')
try:
    nltk.data.find('corpora/wordnet')
except LookupError:
    nltk.download('wordnet')
try:
    nltk.data.find('taggers/averaged_perceptron_tagger')
except LookupError:
    nltk.download('averaged_perceptron_tagger')

# Configuration
API_KEY = # "YOUR_API_KEY_HERE"  # Replace with your actual API key
INPUT_DIR = "./"  # Directory containing PDF files (current directory)
OUTPUT_FILE = "academic_paper_summaries.xlsx"  # Output Excel file
MAX_SUMMARY_WORDS = 200  # Maximum words for the summary, matching llm.py's approach

# Rate limiting configuration
TOKEN_LIMIT_PER_MINUTE = 20000  # Claude's rate limit
WORD_TO_TOKEN_RATIO = 1.75  # Assume 1 word = 1.75 tokens
MAX_DOCUMENT_TOKENS = 20000  # Maximum tokens for document processing

# Initialize Claude client
client = anthropic.Anthropic(api_key=API_KEY)

# Initialize NLTK components
lemmatizer = WordNetLemmatizer()
stop_words = set(stopwords.words('english'))
punct_translator = str.maketrans('', '', string.punctuation)

# Token usage tracker
class TokenUsageTracker:
    def __init__(self, limit_per_minute):
        self.limit = limit_per_minute
        self.used = 0
        self.last_reset = time.time()
    
    def estimate_tokens(self, content):
        """Estimate token count based on word count"""
        word_count = len(content.split())
        return int(word_count * WORD_TO_TOKEN_RATIO)
    
    def can_process(self, content):
        """Check if we can process the content without exceeding rate limits"""
        # Reset counter if more than a minute has passed
        current_time = time.time()
        if current_time - self.last_reset >= 60:
            self.used = 0
            self.last_reset = current_time
        
        # Estimate tokens for this content
        estimated_tokens = self.estimate_tokens(content)
        
        # Check if we have enough quota
        return self.used + estimated_tokens <= self.limit
    
    def record_usage(self, content):
        """Record token usage after processing content"""
        self.used += self.estimate_tokens(content)
    
    def wait_for_reset(self):
        """Wait until the rate limit resets - always wait a full minute"""
        print(f"Rate limit approaching. Waiting for 60 seconds...")
        time.sleep(60)  # Always wait a full minute
        
        # Reset counter
        self.used = 0
        self.last_reset = time.time()

# Initialize token tracker
token_tracker = TokenUsageTracker(TOKEN_LIMIT_PER_MINUTE)

# Function to truncate text to stay within token limit
def truncate_to_token_limit(text, max_tokens=MAX_DOCUMENT_TOKENS):
    """
    Truncate text to stay within token limit
    
    Args:
        text: The text to truncate
        max_tokens: Maximum tokens allowed
        
    Returns:
        Truncated text
    """
    max_words = int(max_tokens / WORD_TO_TOKEN_RATIO)
    words = text.split()
    
    if len(words) <= max_words:
        return text, len(words)
    
    print(f"Truncating document from {len(words)} words to {max_words} words to stay within token limit")
    return " ".join(words[:max_words]), len(words)  # Return truncated text and original word count

# Function to clean text using NLTK instead of spaCy
def clean_text(text):
    """Cleans and lemmatizes the text using NLTK"""
    if not text:
        return text
    
    try:
        # Tokenize text into words
        tokens = word_tokenize(text.lower())
        
        # Filter out stop words, punctuation, and lemmatize
        cleaned_tokens = []
        for token in tokens:
            # Remove punctuation
            token = token.translate(punct_translator)
            
            # Skip empty tokens, stop words, and non-alphabetic tokens
            if (not token or 
                token in stop_words or 
                not token.isalpha()):
                continue
                
            # Lemmatize the token
            lemma = lemmatizer.lemmatize(token)
            cleaned_tokens.append(lemma)
        
        # Join the cleaned tokens back into text
        cleaned_text = " ".join(cleaned_tokens)
        return cleaned_text
    except Exception as e:
        print(f"Error cleaning text: {e}")
        return text

def extract_text_from_pdf(file_path):
    """Extract text from PDF file using multiple methods for better coverage"""
    try:
        # First attempt with pdfplumber (from llm.py)
        text = ""
        try:
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n\n"
            
            # If we got good content, return it
            if len(text.strip()) > 100:
                return text
        except Exception as e:
            print(f"pdfplumber extraction failed: {e}")
        
        # Second attempt with PyPDF2 if pdfplumber didn't give good results
        if len(text.strip()) < 100:
            text = ""
            with open(file_path, "rb") as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page_num in range(len(pdf_reader.pages)):
                    page = pdf_reader.pages[page_num]
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n\n"
        
        return text
    except Exception as e:
        print(f"Error extracting text from PDF: {e}")
        return ""

def extract_metadata_from_academic_paper(text):
    """Extract metadata from academic paper text"""
    metadata = {}
    
    # Extract title - usually in the first few lines
    title_pattern = re.compile(r'^([^\n]{10,150})(?:\n|$)', re.MULTILINE)
    title_match = title_pattern.search(text[:500])
    metadata['title'] = title_match.group(1).strip() if title_match else ""
    
    # Clean up the title
    if metadata['title']:
        # Remove common paper indicators
        metadata['title'] = re.sub(r'^(A|An|The)\s+Survey\s+of\s*:', '', metadata['title'])
        metadata['title'] = re.sub(r'^arXiv:.+?(\s+|$)', '', metadata['title'])
    
    # Extract date - look for a year in the first few pages
    date_pattern = re.compile(r'(?:(?:19|20)\d{2})')
    date_matches = date_pattern.findall(text[:2000])
    metadata['date'] = date_matches[0] if date_matches else ""
    
    # Extract source - usually a journal or conference
    source_pattern = re.compile(r'(?:published\s+in|proceedings\s+of|journal\s+of|conference\s+on)\s+([^.\n]{5,100})', re.IGNORECASE)
    source_match = source_pattern.search(text[:2000])
    metadata['source'] = source_match.group(1).strip() if source_match else ""
    
    # If source is not found, try to extract arXiv ID
    if not metadata['source']:
        arxiv_pattern = re.compile(r'arXiv:(\d{4}\.\d{5})', re.IGNORECASE)
        arxiv_match = arxiv_pattern.search(text[:1000])
        metadata['source'] = f"arXiv:{arxiv_match.group(1)}" if arxiv_match else ""
    
    # Extract authors
    author_pattern = re.compile(r'(?:author[s]?:?\s*)([A-Z][a-z]+(?:\s+[A-Z][a-z]+)+(?:,\s*[A-Z][a-z]+(?:\s+[A-Z][a-z]+)+)*)', re.IGNORECASE)
    author_match = author_pattern.search(text[:2000])
    authors = author_match.group(1).strip() if author_match else ""
    metadata['topic'] = authors  # Storing authors in the topic field for now
    
    # Set URL to blank (academic papers usually don't have URLs in our context)
    metadata['url'] = ""
    
    # Set category to "Academic Paper"
    metadata['category'] = "Academic Paper"
    
    return metadata

def generate_summary(content, metadata):
    """Generate a summary using Claude API - using improved prompt to avoid titles"""
    
    # Modified prompt to explicitly avoid titles and # symbols
    prompt = f"""Summarize the following text by extracting and paraphrasing only the core information and key concepts in a single paragraph, limited to 100 words.
            Strict formatting rules:
            - IMPORTANT: Do NOT generate the summary with any titles or headings
            - Do not include rhetorical questions.
            - Remove all conversational or reflective phrases (e.g., "Let's break down...", "Do you want to...").
            - Do not include any introductory or closing remarks.
            - Write in a neutral, academic tone.
            - Present all key methods, terms, and examples clearly and concisely in paragraph form (no bullet points).
            - Ensure that the summary always ends with a full and complete sentence.
            - Keep the summary short, recommended to be around 100 words.
            - Begin directly with the summary text without any prefixes.

            Document:
            \"\"\"{content}\"\"\"
            """
    
    # Check if we can make this API call without hitting rate limits
    if not token_tracker.can_process(content):
        token_tracker.wait_for_reset()
    
    try:
        response = client.messages.create(
            model="claude-3-7-sonnet-20250219",
            max_tokens=1024,
            temperature=0.0,
            system="You are a professional academic paper summarizer. Create concise, accurate summaries that capture the essential contributions, findings, and conclusions of academic papers. Your summaries should be factual, well-structured, and highlight the key insights and methodological approaches. NEVER start summaries with titles, headings, or # symbols.",
            messages=[
                {"role": "user", "content": prompt}
            ]
        )
        
        # Record token usage
        token_tracker.record_usage(content)
        
        summary = response.content[0].text
        
        # Further clean the summary to remove any remaining '#' symbols or titles at the beginning
        summary = re.sub(r'^#\s*[^\n]*\n', '', summary)  # Remove title with # if present
        summary = re.sub(r'^Title:', '', summary)  # Remove "Title:" prefix if present
        summary = re.sub(r'^Summary:', '', summary)  # Remove "Summary:" prefix if present
        summary = summary.strip()
        
        # Ensure summary is under MAX_SUMMARY_WORDS words
        words = summary.split()
        if len(words) > MAX_SUMMARY_WORDS:
            summary = " ".join(words[:MAX_SUMMARY_WORDS]) + "..."
            
        return summary
        
    except Exception as e:
        print(f"Error generating summary: {e}")
        return f"Error generating summary: {e}"

def count_words(text):
    """Count the number of words in text"""
    return len(text.split())

def read_existing_excel():
    """Read the existing Excel file if it exists"""
    if not os.path.exists(OUTPUT_FILE):
        print(f"Excel file {OUTPUT_FILE} not found. Will create a new one.")
        return None
    
    try:
        df = pd.read_excel(OUTPUT_FILE)
        print(f"Successfully read {len(df)} entries from {OUTPUT_FILE}")
        return df
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

def needs_regeneration(summary):
    """Check if a summary needs to be regenerated"""
    if not isinstance(summary, str):
        return True
        
    summary = summary.strip()
    return (summary.startswith('#') or 
            summary.startswith('Error') or 
            summary.startswith('Title:') or
            summary.startswith('Summary:'))

def process_existing_excel():
    """Process entries in the existing Excel file"""
    # Read the existing Excel file
    df = read_existing_excel()
    if df is None or len(df) == 0:
        print("No existing data to process.")
        return False
    
    # Check if "Document Word Count" column exists, add it if it doesn't
    if 'Document Word Count' not in df.columns:
        df['Document Word Count'] = 0
        print("Added 'Document Word Count' column to the existing spreadsheet.")
    
    # Identify entries with problematic summaries (need regeneration)
    summary_mask = df['Summary'].apply(needs_regeneration)
    summary_count = summary_mask.sum()
    
    # Identify entries with zero word count 
    word_count_mask = df['Document Word Count'] == 0
    word_count_count = word_count_mask.sum()
    
    if summary_count == 0 and word_count_count == 0:
        print("No entries need processing.")
        return False
    
    print(f"Found {summary_count} summaries that need regeneration.")
    print(f"Found {word_count_count} entries with zero word count that need calculation.")
    
    # First, process entries that need only word count calculation (not summary regeneration)
    word_count_only_mask = word_count_mask & ~summary_mask
    word_count_only_count = word_count_only_mask.sum()
    
    if word_count_only_count > 0:
        print(f"Processing {word_count_only_count} entries that need word count calculation only...")
        
        for idx, row in tqdm(df[word_count_only_mask].iterrows(), total=word_count_only_count, desc="Calculating word counts"):
            file_path = os.path.join(INPUT_DIR, row['Filename'])
            
            if not os.path.exists(file_path):
                print(f"File not found: {file_path}. Skipping.")
                continue
            
            try:
                print(f"Calculating word count for: {row['Filename']}")
                
                # Extract text from PDF
                full_text = extract_text_from_pdf(file_path)
                
                if not full_text:
                    print(f"  Warning: Could not extract text from {row['Filename']}")
                    continue
                
                # Calculate and update word count
                original_word_count = count_words(full_text)
                df.at[idx, 'Document Word Count'] = original_word_count
                
                print(f"  Updated word count: {original_word_count} words")
                
                # Save after each update to avoid losing progress
                df.to_excel(OUTPUT_FILE, index=False)
                
            except Exception as e:
                print(f"  Error calculating word count for {row['Filename']}: {e}")
    
    # Next, process entries that need summary regeneration
    if summary_count > 0:
        print(f"Processing {summary_count} entries that need summary regeneration...")
        
        for idx, row in tqdm(df[summary_mask].iterrows(), total=summary_count, desc="Regenerating summaries"):
            file_path = os.path.join(INPUT_DIR, row['Filename'])
            
            if not os.path.exists(file_path):
                print(f"File not found: {file_path}. Skipping.")
                continue
            
            try:
                print(f"Regenerating summary for: {row['Filename']}")
                
                # Extract text from PDF
                full_text = extract_text_from_pdf(file_path)
                
                if not full_text:
                    print(f"  Warning: Could not extract text from {row['Filename']}")
                    continue
                
                # Update word count if needed
                if row['Document Word Count'] == 0:
                    original_word_count = count_words(full_text)
                    df.at[idx, 'Document Word Count'] = original_word_count
                    print(f"  Updated word count: {original_word_count} words")
                
                # Clean the full text using NLTK
                cleaned_text = clean_text(full_text)
                
                # Only use cleaned version if it has sufficient content
                content_for_summary = cleaned_text if len(cleaned_text) > len(full_text) * 0.2 else full_text
                
                # Truncate if needed to stay within token limits
                truncated_content, _ = truncate_to_token_limit(content_for_summary)
                
                # Generate new summary
                new_summary = generate_summary(truncated_content, {})
                
                # Update the DataFrame
                df.at[idx, 'Summary'] = new_summary
                df.at[idx, 'Summary Length'] = count_words(new_summary)
                
                print(f"  Updated summary for {row['Filename']}")
                
                # Save after each update to avoid losing progress
                df.to_excel(OUTPUT_FILE, index=False)
                
            except Exception as e:
                print(f"  Error regenerating summary for {row['Filename']}: {e}")
    
    print("Processing of existing entries completed.")
    return True

def get_new_pdf_files(existing_df):
    """Get list of PDF files not already in the Excel file"""
    # Find all PDF files in the directory
    all_pdf_files = glob.glob(os.path.join(INPUT_DIR, "*.pdf"))
    
    # If there's an existing dataframe, get the filenames already processed
    if existing_df is not None and len(existing_df) > 0:
        existing_filenames = set(existing_df['Filename'].tolist())
    else:
        existing_filenames = set()
    
    # Find new PDF files
    new_pdf_files = []
    for file_path in all_pdf_files:
        filename = os.path.basename(file_path)
        if filename not in existing_filenames:
            new_pdf_files.append(file_path)
    
    return new_pdf_files

def append_to_excel(new_data, existing_df):
    """Append new data to existing Excel file"""
    if existing_df is not None and len(existing_df) > 0:
        # Combine existing and new data
        all_data = pd.concat([existing_df, pd.DataFrame(new_data)], ignore_index=True)
    else:
        # Create new dataframe if no existing data
        all_data = pd.DataFrame(new_data)
    
    # Save to Excel with formatting
    save_to_excel(all_data)

def save_to_excel(df):
    """Save DataFrame to Excel with proper formatting"""
    # Save to Excel with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Paper Summaries"
    
    # Add headers
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            if col_num == headers.index("Summary") + 1:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Auto-adjust column width
    for column in range(1, len(headers) + 1):
        if column == headers.index("Summary") + 1:
            # Make summary column wider
            ws.column_dimensions[get_column_letter(column)].width = 60
        else:
            # Auto-adjust other columns based on content
            max_length = 0
            for row in range(1, len(df) + 2):
                cell_value = str(ws.cell(row=row, column=column).value or "")
                max_length = max(max_length, len(cell_value))
            ws.column_dimensions[get_column_letter(column)].width = min(max_length + 2, 30)
    
    # Freeze the header row
    ws.freeze_panes = "A2"
    
    # Save the workbook
    wb.save(OUTPUT_FILE)
    print(f"\nSummaries saved to {OUTPUT_FILE}")

def process_files():
    """Process all PDF files and create/update Excel spreadsheet"""
    
    # First, process any existing Excel file entries
    if process_existing_excel():
        print("Processing of existing entries completed.")
    
    # Read the existing Excel file to get list of already processed files
    existing_df = read_existing_excel()
    
    # Get only the new PDF files that haven't been processed yet
    new_pdf_files = get_new_pdf_files(existing_df)
    
    if not new_pdf_files:
        print("\nNo new PDF files found to process.")
        return
    
    print(f"\nFound {len(new_pdf_files)} new PDF files to process.")
    
    # Prepare data for new files
    new_data = []
    
    for i, file_path in enumerate(tqdm(new_pdf_files, desc="Processing new PDFs")):
        filename = os.path.basename(file_path)
        print(f"[{i+1}/{len(new_pdf_files)}] Processing: {filename}")
        
        try:
            # Extract text from PDF
            full_text = extract_text_from_pdf(file_path)
            
            if not full_text:
                print(f"  Warning: Could not extract text from {filename}")
                continue
            
            # Get original word count
            original_word_count = count_words(full_text)
            
            # Extract metadata
            metadata = extract_metadata_from_academic_paper(full_text)
            
            # Clean the full text using NLTK
            cleaned_text = clean_text(full_text)
            
            # Only use cleaned version if it has sufficient content
            content_for_summary = cleaned_text if len(cleaned_text) > len(full_text) * 0.2 else full_text
            
            # Truncate if needed to stay within token limits
            truncated_content, _ = truncate_to_token_limit(content_for_summary)
            
            # Log extracted content length
            content_word_count = count_words(truncated_content)
            print(f"  Extracted {content_word_count} words from document (original: {original_word_count})")
            
            # Generate summary
            print(f"  Generating summary for: {metadata['title']}")
            summary = generate_summary(truncated_content, metadata)
            
            # Calculate summary length
            summary_length = count_words(summary)
            print(f"  Generated summary of {summary_length} words")
            
            # Format date if possible
            date_str = metadata['date']
            formatted_date = date_str
            
            # Add to new data
            new_data.append({
                'Filename': filename,
                'Title': metadata['title'],
                'Date': formatted_date,
                'Source': metadata['source'],
                'Topic': metadata['topic'],  # Contains authors
                'Category': metadata['category'],
                'Summary': summary,
                'Summary Length': summary_length,
                'Document Word Count': original_word_count,  # Add original word count
                'URL': metadata['url']
            })
            
        except Exception as e:
            print(f"  Error processing {filename}: {e}")
            import traceback
            traceback.print_exc()
            
    # Append new data to existing Excel file
    if new_data:
        append_to_excel(new_data, existing_df)
        print(f"\nCompleted! Added {len(new_data)} new summaries to {OUTPUT_FILE}")
    else:
        print("\nNo new data to save.")

if __name__ == "__main__":
    print("Starting academic paper summarization process...")
    process_files()
    print("Process completed!")
